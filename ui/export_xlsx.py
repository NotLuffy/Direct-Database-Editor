"""
CNC Direct Editor — Excel (.xlsx) exporter.

Generates a multi-sheet workbook:
  • One sheet per round size  (used files + FREE rows for unused O-numbers)
  • One "All" sheet           (all sheets combined, same structure)

Each sheet:
  - Sorted by Part Type → CB (mm) → Thickness
  - Autofilter on every column
  - Frozen header row
  - FREE rows at the bottom (O-number populated, all other cells "FREE")
  - N/A in cells that don't apply to the file type
"""

import os
import re
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

import direct_database as db
from direct_models import _part_type
from verifier import parse_title_specs

# ---------------------------------------------------------------------------
# Round-size → O-number range table  (mirrors verifier._ROUND_TO_O_RANGE)
# ---------------------------------------------------------------------------

_ROUND_SHEETS = [
    # (sheet_name,          round_sizes (in),      o_min,  o_max )
    ("5.75in",              [5.75],                50000,  59999),
    ("6.00in",              [6.00],                60000,  62499),
    ("6.25in",              [6.25],                62500,  64999),
    ("6.50in",              [6.50],                65000,  69999),
    ("7.00in",              [7.00],                70000,  74999),
    ("7.50in",              [7.50],                75000,  79999),
    ("8.00in",              [8.00],                80000,  84999),
    ("8.50in",              [8.50],                85000,  89999),
    ("9.50in",              [9.50],                90000,  99999),
    ("10.25-10.50in",       [10.25, 10.50],        10000,  10999),
    ("13.00in",             [13.00],               13000,  13999),
]

_HEADERS = [
    "O-Number",
    "Title",
    "Round Size",
    "CB (mm)",
    "OB (mm)",
    "Thickness",
    "Hub",
    "Type",
    "Notes",
    "Verify Status",
    "Fails",
]

_COL_WIDTHS = [12, 44, 12, 10, 10, 12, 10, 12, 32, 38, 26]

# ---------------------------------------------------------------------------
# Palette  (light / standard Excel look)
# ---------------------------------------------------------------------------

_HDR_FILL  = PatternFill("solid", fgColor="2F5496")   # deep blue header
_HDR_FONT  = Font(bold=True, color="FFFFFF", name="Calibri", size=10)
_ROW_FONT  = Font(color="000000", name="Calibri", size=10)
_FREE_FILL = PatternFill("solid", fgColor="F2F2F2")   # light grey stripe
_FREE_FONT = Font(color="AAAAAA", name="Calibri", size=10, italic=True)
_FREE_ONUM = Font(color="888888", name="Calibri", size=10, italic=True)
_FAIL_FONT = Font(color="C00000", name="Calibri", size=10, bold=True)

_HDR_ALIGN = Alignment(horizontal="center", vertical="center")
_CTR_ALIGN = Alignment(horizontal="center", vertical="center")

# ---------------------------------------------------------------------------
# Spec helpers
# ---------------------------------------------------------------------------

def _specs(title: str) -> dict:
    if not title:
        return {}
    try:
        return parse_title_specs(title) or {}
    except Exception:
        return {}


def _fmt_mm(v) -> str:
    return f"{v:.1f}" if v is not None else "N/A"


def _fmt_in(v) -> str:
    return f'{v:.3f}"' if v is not None else "N/A"


def _fmt_hub(hc_in) -> str:
    if hc_in is None:
        return "N/A"
    mm = hc_in * 25.4
    if abs(mm - 15.0) < 0.15:
        return '15MM (0.591")'
    return f'{hc_in:.3f}"'


def _fails(vstatus: str) -> str:
    if not vstatus:
        return ""
    return " ".join(t for t in vstatus.split() if t.upper().endswith(":FAIL"))


def _o_int(o_number: str) -> int:
    m = re.match(r'O(\d+)', o_number or "", re.IGNORECASE)
    return int(m.group(1)) if m else 0

# ---------------------------------------------------------------------------
# Type sort order
# ---------------------------------------------------------------------------

_ONUM_FNAME_RE = re.compile(r'^O(\d{4,6})(?:_\d+)?', re.IGNORECASE)


def _rs_label_for_o(o_int: int) -> str:
    """Return the round-size display string for an O-number, e.g. '5.75"'."""
    for _name, round_sizes, o_min, o_max in _ROUND_SHEETS:
        if o_min <= o_int <= o_max:
            return "/".join(f'{rs:.2f}"' for rs in round_sizes)
    return ""


def _scan_folder_o_ints(folder: str) -> set[int]:
    """Walk a folder on disk and return every O-number integer found in filenames.
    Used to catch files that are on disk but not yet in the DB.
    """
    result: set[int] = set()
    try:
        for root, dirs, files in os.walk(folder):
            dirs[:] = [d for d in dirs if not d.startswith(".")]
            for fname in files:
                base = os.path.splitext(fname)[0]
                m = _ONUM_FNAME_RE.match(base)
                if m:
                    try:
                        result.add(int(m.group(1)))
                    except ValueError:
                        pass
    except Exception:
        pass
    return result


_TYPE_ORDER = {
    "STD": 0, "HC": 1, "15MM HC": 2, "2PC": 3,
    "STEP": 4, "STEEL": 5, "SPACER": 6, "LUG": 7, "STUD": 8,
}


def _sort_key(r: dict) -> tuple:
    """Sort: round_size → type → CB → thickness."""
    return (
        r["_rs"],
        _TYPE_ORDER.get(r["_type"], 9),
        r["_cb"],
        r["_th"],
    )

# ---------------------------------------------------------------------------
# Build enriched row dict from a DB record
# ---------------------------------------------------------------------------

def _build_row(rec) -> dict:
    title  = rec["program_title"] or ""
    sp     = _specs(title)
    ptype  = _part_type(title)

    rs_in  = sp.get("round_size_in")
    cb_mm  = sp.get("cb_mm")
    ob_mm  = sp.get("ob_mm")
    th_in  = sp.get("length_in")
    hc_in  = sp.get("hc_height_in")

    vstatus = rec["verify_status"] or ""

    th_disp = f"{th_in * 25.4:.1f}MM" if th_in is not None else "N/A"

    return {
        # Display columns (in header order)
        "o_number":  rec["o_number"] or "",
        "title":     title,
        "rs_disp":   f'{rs_in:.2f}"' if rs_in else "N/A",
        "cb":        _fmt_mm(cb_mm),
        "ob":        _fmt_mm(ob_mm),         # N/A when ob_mm is None
        "thickness": th_disp,
        "hub":       _fmt_hub(hc_in),
        "type":      ptype,
        "notes":     (rec["notes"] or "").replace("\n", " ")[:200],
        "verify":    vstatus,
        "fails":     _fails(vstatus),
        # Sort/filter keys (not written to sheet)
        "_rs":   rs_in or 0.0,
        "_cb":   cb_mm if cb_mm is not None else 99999.0,
        "_th":   th_in if th_in is not None else 99999.0,
        "_type": ptype,
        "_onum": _o_int(rec["o_number"] or ""),
    }

# ---------------------------------------------------------------------------
# Write one worksheet
# ---------------------------------------------------------------------------

_RS_COL = 3   # 1-based index of "Round Size" in _HEADERS


def _write_sheet(ws, used_rows: list[dict], free_onums: list[int],
                 sort_by_onum: bool = False,
                 free_rs_label: str | None = None) -> None:
    """
    Write header + used rows (sorted) + FREE rows to *ws*.

    sort_by_onum  : sort by O-number ascending (True for the All sheet).
    free_rs_label : fixed round-size label for all FREE rows on this sheet.
                    When None (All sheet) the label is looked up per O-number.
    """

    # ── Header ───────────────────────────────────────────────────────────
    for col_i, (hdr, width) in enumerate(zip(_HEADERS, _COL_WIDTHS), start=1):
        cell = ws.cell(row=1, column=col_i, value=hdr)
        cell.fill = _HDR_FILL
        cell.font = _HDR_FONT
        cell.alignment = _HDR_ALIGN
        ws.column_dimensions[get_column_letter(col_i)].width = width
    ws.row_dimensions[1].height = 18

    # ── Used rows ────────────────────────────────────────────────────────
    if sort_by_onum:
        sorted_rows = sorted(used_rows, key=lambda r: r["_onum"])
    else:
        sorted_rows = sorted(used_rows, key=_sort_key)

    row_i = 2
    for r in sorted_rows:
        vals = [
            r["o_number"], r["title"],    r["rs_disp"],
            r["cb"],       r["ob"],       r["thickness"],
            r["hub"],      r["type"],     r["notes"],
            r["verify"],   r["fails"],
        ]
        for col_i, val in enumerate(vals, start=1):
            cell = ws.cell(row=row_i, column=col_i, value=val)
            # Fails column (11) in bold red if non-empty
            if col_i == 11 and val:
                cell.font = _FAIL_FONT
            else:
                cell.font = _ROW_FONT
        row_i += 1

    # ── FREE rows ────────────────────────────────────────────────────────
    for o_int in sorted(free_onums):
        o_str = f"O{o_int:05d}"
        rs    = free_rs_label if free_rs_label is not None else _rs_label_for_o(o_int)
        for col_i in range(1, len(_HEADERS) + 1):
            if col_i == 1:
                val = o_str
            elif col_i == _RS_COL:
                val = rs          # round size always populated for FREE rows
            else:
                val = "FREE"
            cell = ws.cell(row=row_i, column=col_i, value=val)
            cell.fill = _FREE_FILL
            cell.font = _FREE_ONUM if col_i == 1 else _FREE_FONT
        row_i += 1

    # ── Autofilter + freeze ───────────────────────────────────────────────
    last_col = get_column_letter(len(_HEADERS))
    ws.auto_filter.ref = f"A1:{last_col}{max(row_i - 1, 1)}"
    ws.freeze_panes = "A2"

# ---------------------------------------------------------------------------
# Public entry point
# ---------------------------------------------------------------------------

def export_workbook(db_path: str, out_path: str,
                    scan_folders: list[str] | None = None) -> tuple[int, int]:
    """
    Build and save the multi-sheet .xlsx workbook.

    scan_folders: when provided, the export also walks these directories on disk
                  to catch files that exist on disk but haven't been imported into
                  the DB yet — they still count as "used" and won't show as FREE.

    Returns (used_count, total_free_count).
    """
    conn = db.get_connection(db_path)
    # Display rows: files that have been scanned (last_seen IS NOT NULL)
    db_rows = conn.execute(
        "SELECT o_number, program_title, verify_status, notes "
        "FROM files WHERE last_seen IS NOT NULL ORDER BY o_number"
    ).fetchall()
    # Used set: ALL files in DB — so anything imported (even without rescan)
    # does not appear as FREE
    all_o_rows = conn.execute(
        "SELECT o_number FROM files WHERE o_number IS NOT NULL"
    ).fetchall()
    conn.close()

    # Build enriched row dicts for every scanned file
    all_built: list[dict] = [_build_row(r) for r in db_rows]

    # Start with every O-number in the DB
    used_o_ints: set[int] = {_o_int(r["o_number"]) for r in all_o_rows}
    used_o_ints.discard(0)

    # Also walk scan_folders on disk — catches files present on disk but not yet
    # imported into the DB (e.g. newly added .nc files not yet rescanned)
    if scan_folders:
        for folder in scan_folders:
            used_o_ints |= _scan_folder_o_ints(folder)

    wb = openpyxl.Workbook()
    wb.remove(wb.active)   # drop the default blank sheet

    # ── "All" sheet (first tab) ───────────────────────────────────────────
    all_free: list[int] = []
    for _, _, o_min, o_max in _ROUND_SHEETS:
        all_free.extend(o for o in range(o_min, o_max + 1) if o not in used_o_ints)

    ws_all = wb.create_sheet(title="All")
    # All sheet: free_rs_label=None → looked up per O-number
    _write_sheet(ws_all, all_built, all_free, sort_by_onum=True)

    # ── Per-round-size sheets ─────────────────────────────────────────────
    for sheet_name, round_sizes, o_min, o_max in _ROUND_SHEETS:
        sheet_rows = [
            r for r in all_built
            if any(abs(r["_rs"] - rs) < 0.01 for rs in round_sizes)
        ]
        free_nums = [
            o for o in range(o_min, o_max + 1) if o not in used_o_ints
        ]
        # Fixed label for all FREE rows on this per-round sheet
        rs_label = "/".join(f'{rs:.2f}"' for rs in round_sizes)
        ws = wb.create_sheet(title=sheet_name)
        _write_sheet(ws, sheet_rows, free_nums,
                     sort_by_onum=False, free_rs_label=rs_label)

    wb.save(out_path)
    return len(all_built), len(all_free)


# ---------------------------------------------------------------------------
# Daily report — files created on a specific date
# ---------------------------------------------------------------------------

def export_daily_report(db_path: str, out_path: str, date_str: str) -> int:
    """
    Export a single-sheet workbook listing files created on *date_str*.

    date_str: YYYY-MM-DD.
    Returns the number of rows written.
    """
    rows = db.get_files_by_index_date(db_path, date_str)
    if not rows:
        return 0

    built = [_build_row(r) for r in rows]

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = date_str   # sheet named after the date

    _write_sheet(ws, built, [], sort_by_onum=True)

    wb.save(out_path)
    return len(built)
